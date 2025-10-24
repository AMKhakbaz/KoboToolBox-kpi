from __future__ import annotations

import csv
import io
import json
from datetime import date, timedelta
from collections.abc import Iterable
from typing import Any, Iterator

from django.contrib.auth import get_user_model
from django.db import ProgrammingError, transaction
from django.db.models import Avg, Count, DurationField, ExpressionWrapper, F, Q
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse, QueryDict, StreamingHttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.encoding import smart_str
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from rest_framework import mixins, serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.views import APIView

from .models import (
    DialerAssignment,
    InsightMembership,
    InsightProject,
    Interview,
    QuotaCell,
    QuotaScheme,
    SampleContact,
    archive_project,
    build_sample_pool_for_cell,
    deactivate_single_membership,
)
from .pagination import InsightZenPagination
from .serializers import (
    AssignmentFailSerializer,
    AssignmentStatusSerializer,
    DialerAssignmentSerializer,
    InsightMembershipSerializer,
    InsightMembershipSyncSerializer,
    InsightProjectSerializer,
    InsightUserSerializer,
    InterviewActionSerializer,
    InterviewSerializer,
    QuotaCellBulkUpsertSerializer,
    QuotaCellSerializer,
    QuotaCellUpdateSerializer,
    QuotaSchemePublishSerializer,
    QuotaSchemeSerializer,
    QuotaSchemeStatsSerializer,
    SampleContactSerializer,
)

User = get_user_model()


class Echo:
    def write(self, value: str) -> str:  # pragma: no cover - simple passthrough for csv.writer
        return value


def _format_export_value(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, (list, tuple)):
        return ', '.join(smart_str(item) for item in value)
    if isinstance(value, bool):
        return 'true' if value else 'false'
    return smart_str(value)


def _serialize_queryset(
    serializer_class: type[serializers.Serializer],
    queryset: Iterable[Any],
    context: dict[str, Any],
) -> Iterator[dict[str, Any]]:
    for instance in queryset:
        yield serializer_class(instance, context=context).data


def stream_csv_response(
    filename: str,
    columns: list[str],
    rows: Iterator[dict[str, Any]],
) -> StreamingHttpResponse:
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer)

    def row_generator() -> Iterator[str]:
        yield writer.writerow(columns)
        for row in rows:
            yield writer.writerow([_format_export_value(row.get(column, '')) for column in columns])

    response = StreamingHttpResponse(row_generator(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_xlsx_response(
    filename: str,
    sheet_name: str,
    columns: list[str],
    rows: Iterator[dict[str, Any]],
) -> HttpResponse:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(columns)
    for row in rows:
        sheet.append([_format_export_value(row.get(column, '')) for column in columns])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _get_accessible_project_ids(user: User) -> Iterable[int]:
    if user.is_superuser or user.is_staff:
        return InsightProject.objects.values_list('id', flat=True)
    return InsightMembership.objects.filter(user=user, is_active=True).values_list('project_id', flat=True)


TELEPHONE_PANEL_KEY = 'telephone-interviewer'
QUOTA_PANEL_KEY = 'quota-management'
COLLECTION_PERFORMANCE_PANEL_KEY = 'collection-performance'
MANAGER_ROLES = {'admin', 'manager'}


def _get_memberships_by_project(user: User) -> dict[int, InsightMembership]:
    return {
        membership.project_id: membership
        for membership in InsightMembership.objects.filter(user=user, is_active=True)
    }


def _panel_accessible_projects(user: User, panel_key: str) -> set[int]:
    if user.is_superuser or user.is_staff:
        return set(InsightProject.objects.values_list('id', flat=True))

    allowed: set[int] = set()
    for membership in InsightMembership.objects.filter(user=user, is_active=True):
        if membership.role in MANAGER_ROLES:
            allowed.add(membership.project_id)
            continue
        permissions = membership.panel_permissions or {}
        collection_permissions = permissions.get('collection') if isinstance(permissions, dict) else None
        if isinstance(collection_permissions, dict) and collection_permissions.get(panel_key):
            allowed.add(membership.project_id)
    return allowed


def _ensure_panel_permission(user: User, project_id: int, panel_key: str) -> InsightMembership | None:
    if user.is_superuser or user.is_staff:
        return None

    try:
        membership = InsightMembership.objects.get(user=user, project_id=project_id, is_active=True)
    except InsightMembership.DoesNotExist as exc:
        raise PermissionDenied(_('You do not have access to this project.')) from exc

    if membership.role in MANAGER_ROLES:
        return membership

    permissions = membership.panel_permissions or {}
    collection_permissions = permissions.get('collection') if isinstance(permissions, dict) else None
    has_access = isinstance(collection_permissions, dict) and collection_permissions.get(panel_key)
    if not has_access:
        raise PermissionDenied(_('You do not have permission to access this panel.'))
    return membership


def _parse_date_param(value: str | None, *, param_name: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:  # pragma: no cover - defensive branch
        raise ValidationError({param_name: _('Enter a valid date in YYYY-MM-DD format.')}) from exc


def _resolve_date_range(params) -> tuple[date, date]:
    date_to = _parse_date_param(params.get('to'), param_name='to') or timezone.now().date()
    default_start = date_to - timedelta(days=6)
    date_from = _parse_date_param(params.get('from'), param_name='from') or default_start
    if date_from > date_to:
        raise ValidationError({'from': _('The start date must not be after the end date.')})
    return date_from, date_to


def _parse_multi_int(params, key: str) -> list[int]:
    raw_values = params.getlist(key)
    result: list[int] = []
    for raw in raw_values:
        if raw is None:
            continue
        for part in str(raw).split(','):
            part = part.strip()
            if not part:
                continue
            try:
                result.append(int(part))
            except ValueError as exc:
                raise ValidationError({key: _('Enter numeric identifiers.')}) from exc
    return result


def _parse_multi_str(params, key: str) -> list[str]:
    raw_values = params.getlist(key)
    result: list[str] = []
    for raw in raw_values:
        if raw is None:
            continue
        for part in str(raw).split(','):
            value = part.strip()
            if value:
                result.append(value)
    return result


def _mask_phone_number(phone_number: str | None) -> str:
    if not phone_number:
        return ''
    digits = phone_number.strip()
    if len(digits) <= 4:
        return digits
    visible = digits[:-3]
    return f"{visible}{'*' * 3}"


def _duration_seconds(interview: Interview) -> int | None:
    if not interview.start_form or not interview.end_form:
        return None
    delta = interview.end_form - interview.start_form
    seconds = int(delta.total_seconds())
    return max(seconds, 0)


def _interviewer_display_name(user: User) -> str:
    profile = getattr(user, 'insight_profile', None)
    if profile and hasattr(profile, 'resolved_display_name'):
        return profile.resolved_display_name
    full_name = user.get_full_name().strip()
    return full_name or user.username


def _resolve_interviewer_label_map(interviewer_ids: Iterable[int]) -> dict[int, str]:
    if not interviewer_ids:
        return {}
    users = (
        User.objects.filter(pk__in=set(interviewer_ids))
        .select_related('insight_profile')
    )
    return {user.pk: _interviewer_display_name(user) for user in users}


def _aggregate_by_interviewer(queryset) -> list[dict[str, Any]]:
    duration_expr = ExpressionWrapper(
        F('end_form') - F('start_form'),
        output_field=DurationField(),
    )
    aggregated = (
        queryset.annotate(duration=duration_expr)
        .values('assignment__interviewer_id')
        .annotate(
            attempts=Count('id'),
            completes=Count('id', filter=Q(outcome_code='COMP')),
            avg_duration=Avg('duration', filter=Q(end_form__isnull=False, start_form__isnull=False)),
        )
    )
    interviewer_ids = [row['assignment__interviewer_id'] for row in aggregated]
    label_map = _resolve_interviewer_label_map(interviewer_ids)
    results: list[dict[str, Any]] = []
    for row in aggregated:
        interviewer_id = row['assignment__interviewer_id']
        avg_duration_td = row.get('avg_duration')
        avg_duration_sec = int(avg_duration_td.total_seconds()) if avg_duration_td else 0
        attempts = int(row.get('attempts') or 0)
        completes = int(row.get('completes') or 0)
        success_rate = (completes / attempts) if attempts else 0.0
        results.append(
            {
                'interviewer_id': interviewer_id,
                'label': label_map.get(interviewer_id, str(interviewer_id)),
                'attempts': attempts,
                'completes': completes,
                'sr': success_rate,
                'avg_duration_sec': avg_duration_sec,
            }
        )
    return results


def _calculate_collection_summary(
    queryset,
    *,
    project_id: int,
    date_from: date,
    date_to: date,
):
    duration_expr = ExpressionWrapper(
        F('end_form') - F('start_form'),
        output_field=DurationField(),
    )
    duration_queryset = queryset.annotate(duration=duration_expr)
    aggregates = duration_queryset.aggregate(
        attempts=Count('id'),
        completes=Count('id', filter=Q(outcome_code='COMP')),
        avg_duration=Avg('duration', filter=Q(end_form__isnull=False, start_form__isnull=False)),
    )

    attempts = int(aggregates.get('attempts') or 0)
    completes = int(aggregates.get('completes') or 0)
    avg_duration_td = aggregates.get('avg_duration')
    avg_duration_sec = int(avg_duration_td.total_seconds()) if avg_duration_td else 0
    success_rate = (completes / attempts) if attempts else 0.0

    daily_queryset = queryset.annotate(
        day=TruncDate(Coalesce('end_form', 'start_form')),
    ).values('day').annotate(
        attempts=Count('id'),
        completes=Count('id', filter=Q(outcome_code='COMP')),
    ).order_by('day')

    by_day = []
    for entry in daily_queryset:
        day_value = entry['day']
        day_attempts = int(entry.get('attempts') or 0)
        day_completes = int(entry.get('completes') or 0)
        day_sr = (day_completes / day_attempts) if day_attempts else 0.0
        by_day.append(
            {
                'day': day_value.isoformat() if day_value else None,
                'attempts': day_attempts,
                'completes': day_completes,
                'sr': day_sr,
            }
        )

    return {
        'project': project_id,
        'range': {'from': date_from.isoformat(), 'to': date_to.isoformat()},
        'totals': {
            'attempts': attempts,
            'completes': completes,
            'success_rate': success_rate,
            'avg_duration_sec': avg_duration_sec,
        },
        'by_day': by_day,
    }


def _collection_performance_queryset(
    *,
    user: User,
    project_id: int,
    date_from: date,
    date_to: date,
    interviewer_ids: list[int],
    outcome_codes: list[str],
    teams: list[str],
):
    accessible_projects = _panel_accessible_projects(user, COLLECTION_PERFORMANCE_PANEL_KEY)
    if project_id not in accessible_projects:
        raise PermissionDenied(_('You do not have permission to access this project.'))

    queryset = (
        Interview.objects.select_related(
            'assignment__project',
            'assignment__interviewer__insight_profile',
            'assignment__sample',
            'assignment__cell',
        )
        .filter(assignment__project_id=project_id)
        .annotate(effective_timestamp=Coalesce('end_form', 'start_form'))
        .filter(effective_timestamp__isnull=False)
        .filter(effective_timestamp__date__gte=date_from, effective_timestamp__date__lte=date_to)
    )

    if interviewer_ids:
        queryset = queryset.filter(assignment__interviewer_id__in=interviewer_ids)
    if outcome_codes:
        queryset = queryset.filter(outcome_code__in=outcome_codes)
    if teams:
        queryset = queryset.filter(assignment__interviewer__insight_profile__team__in=teams)

    return queryset


def _resolve_collection_filters(request) -> dict[str, Any]:
    params = request.query_params.copy()
    if request.method != 'GET' and hasattr(request, 'data'):
        data = request.data
        if isinstance(data, QueryDict):
            for key in data:
                for value in data.getlist(key):
                    params.appendlist(key, value)
        elif isinstance(data, dict):
            for key, value in data.items():
                if value is None:
                    continue
                if isinstance(value, (list, tuple)):
                    for item in value:
                        params.appendlist(key, item)
                else:
                    params.appendlist(key, value)
    project_raw = params.get('project')
    if not project_raw:
        raise ValidationError({'project': _('Project is required.')})
    try:
        project_id = int(project_raw)
    except ValueError as exc:
        raise ValidationError({'project': _('Project must be a numeric identifier.')}) from exc

    date_from, date_to = _resolve_date_range(params)
    interviewer_ids = _parse_multi_int(params, 'interviewer')
    outcome_codes = _parse_multi_str(params, 'outcomes')
    teams = _parse_multi_str(params, 'team')

    return {
        'project_id': project_id,
        'date_from': date_from,
        'date_to': date_to,
        'interviewer_ids': interviewer_ids,
        'outcome_codes': outcome_codes,
        'teams': teams,
    }


class CollectionPerformanceBaseView(APIView):
    permission_classes = (IsAuthenticated,)

    def get_filters_and_queryset(self, request):
        filters = _resolve_collection_filters(request)
        queryset = _collection_performance_queryset(user=request.user, **filters)
        return filters, queryset


class InsightUserViewSet(viewsets.ModelViewSet):
    serializer_class = InsightUserSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = InsightZenPagination

    def get_queryset(self):
        queryset = (
            User.objects.all()
            .select_related('insight_profile')
            .prefetch_related('insight_memberships__project')
        )
        if self.request.user.is_staff or self.request.user.is_superuser:
            filtered = queryset
        else:
            accessible_project_ids = list(_get_accessible_project_ids(self.request.user))
            if not accessible_project_ids:
                return queryset.none()
            filtered = queryset.filter(
                Q(pk=self.request.user.pk)
                | Q(insight_memberships__project_id__in=accessible_project_ids)
            )
        return self._apply_user_filters(filtered).distinct()

    def _apply_user_filters(self, queryset):
        params = self.request.query_params
        search = params.get('q')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
                | Q(insight_profile__phone__icontains=search)
            )
        is_active = params.get('is_active')
        if is_active in {'true', 'false'}:
            queryset = queryset.filter(is_active=is_active == 'true')
        role = params.get('role')
        if role:
            queryset = queryset.filter(insight_memberships__role=role)
        project_id = params.get('project_id')
        if project_id:
            queryset = queryset.filter(insight_memberships__project_id=project_id)
        return queryset

    def perform_destroy(self, instance: User) -> None:
        instance.is_active = False
        instance.save(update_fields=['is_active'])

    @action(detail=True, methods=['get'], url_path='memberships')
    def list_memberships(self, request, pk=None):
        user = self.get_object()
        memberships = user.insight_memberships.select_related('project').all()
        serializer = InsightMembershipSerializer(memberships, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='memberships')
    def create_membership(self, request, pk=None):
        user = self.get_object()
        serializer = InsightMembershipSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        membership = serializer.save(user=user)
        return Response(InsightMembershipSerializer(membership, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch', 'delete'], url_path='memberships/(?P<membership_id>[^/.]+)')
    def update_membership(self, request, pk=None, membership_id=None):
        user = self.get_object()
        try:
            membership = user.insight_memberships.get(pk=membership_id)
        except InsightMembership.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.method.lower() == 'delete':
            deactivate_single_membership(membership)
            return Response(status=status.HTTP_204_NO_CONTENT)

        serializer = InsightMembershipSerializer(
            membership, data=request.data, partial=True, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export')
    def export_users(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        columns_param = request.query_params.get('columns')
        columns = [col.strip() for col in columns_param.split(',') if col.strip()] if columns_param else []
        if not columns:
            columns = ['username', 'first_name', 'last_name', 'email', 'phone', 'is_active']
        export_format = request.query_params.get('format', 'csv').lower()
        if export_format not in {'csv', 'xlsx'}:
            return Response({'detail': _('Unsupported export format.')}, status=status.HTTP_400_BAD_REQUEST)

        queryset_iterator = queryset.iterator()

        if export_format == 'csv':
            rows = _serialize_queryset(
                InsightUserSerializer,
                queryset_iterator,
                context=self.get_serializer_context(),
            )
            return stream_csv_response('insightzen-users.csv', columns, rows)

        rows = _serialize_queryset(
            InsightUserSerializer,
            queryset.iterator(),
            context=self.get_serializer_context(),
        )
        return build_xlsx_response('insightzen-users.xlsx', 'Users', columns, rows)


class InsightProjectViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.ListModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = InsightProjectSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = InsightZenPagination

    def get_queryset(self):
        queryset = InsightProject.objects.all().prefetch_related('memberships__user__insight_profile')
        if self.request.user.is_staff or self.request.user.is_superuser:
            filtered = queryset
        else:
            accessible_project_ids = list(_get_accessible_project_ids(self.request.user))
            if not accessible_project_ids:
                return queryset.none()
            filtered = queryset.filter(pk__in=accessible_project_ids)
        return self._apply_project_filters(filtered)

    def _apply_project_filters(self, queryset):
        params = self.request.query_params
        search = params.get('q')
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(code__icontains=search))
        status_param = params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        owner_id = params.get('owner_id')
        if owner_id:
            queryset = queryset.filter(owner_id=owner_id)
        type_param = params.get('type')
        if type_param:
            queryset = queryset.filter(types__icontains=type_param)
        return queryset

    def perform_destroy(self, instance: InsightProject) -> None:
        archive_project(instance)

    @action(detail=True, methods=['get', 'post'], url_path='memberships')
    def project_memberships(self, request, pk=None):
        project = self.get_object()
        if request.method.lower() == 'get':
            memberships = project.memberships.select_related('user').all()
            serializer = InsightMembershipSerializer(memberships, many=True, context={'request': request})
            return Response(serializer.data)

        serializer = InsightMembershipSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        membership = serializer.save(project=project)
        return Response(InsightMembershipSerializer(membership, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['patch', 'delete'], url_path='memberships/(?P<membership_id>[^/.]+)')
    def update_project_membership(self, request, pk=None, membership_id=None):
        project = self.get_object()
        try:
            membership = project.memberships.get(pk=membership_id)
        except InsightMembership.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if request.method.lower() == 'delete':
            deactivate_single_membership(membership)
            return Response(status=status.HTTP_204_NO_CONTENT)

        serializer = InsightMembershipSerializer(
            membership, data=request.data, partial=True, context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='sync-memberships')
    def sync_memberships(self, request, pk=None):
        project = self.get_object()
        payload = request.data.get('memberships', request.data)
        if not isinstance(payload, list):
            raise ValidationError({'memberships': _('A list of memberships is required.')})

        serializer = InsightMembershipSyncSerializer(
            data=payload,
            many=True,
            context={'request': request, 'project': project},
        )
        serializer.is_valid(raise_exception=True)
        memberships_data = serializer.validated_data

        seen_user_ids: set[int] = set()
        for membership_data in memberships_data:
            user_id = membership_data['user'].id
            if user_id in seen_user_ids:
                raise ValidationError({'memberships': _('Each user can only appear once.')})
            seen_user_ids.add(user_id)

        with transaction.atomic():
            existing_memberships = {
                membership.id: membership
                for membership in project.memberships.select_for_update().select_related('user')
            }
            memberships_by_user = {membership.user_id: membership for membership in existing_memberships.values()}
            processed_ids: set[int] = set()

            for membership_data in memberships_data:
                membership_id = membership_data.get('id')
                user = membership_data['user']
                title = membership_data.get('title', '')
                role = membership_data['role']
                panel_permissions = membership_data.get('panel_permissions') or {}

                membership = None
                if membership_id:
                    membership = existing_memberships.get(membership_id)
                    if membership is None:
                        raise ValidationError({'memberships': _('Membership not found.')})
                else:
                    membership = memberships_by_user.get(user.id)

                if membership:
                    membership.user = user
                    membership.title = title
                    membership.role = role
                    membership.panel_permissions = panel_permissions
                    membership.is_active = True
                    membership.save()
                else:
                    membership = InsightMembership.objects.create(
                        project=project,
                        user=user,
                        title=title,
                        role=role,
                        panel_permissions=panel_permissions,
                        is_active=True,
                    )

                processed_ids.add(membership.id)
                memberships_by_user[user.id] = membership

            for membership in existing_memberships.values():
                if membership.id not in processed_ids and membership.is_active:
                    membership.is_active = False
                    membership.save(update_fields=['is_active'])

        refreshed = project.memberships.filter(is_active=True).select_related('user')
        response_serializer = InsightMembershipSerializer(refreshed, many=True, context={'request': request})
        return Response(response_serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export')
    def export_projects(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        columns_param = request.query_params.get('columns')
        columns = [col.strip() for col in columns_param.split(',') if col.strip()] if columns_param else []
        if not columns:
            columns = ['code', 'name', 'status', 'owner_username', 'types']
        export_format = request.query_params.get('format', 'csv').lower()
        if export_format not in {'csv', 'xlsx'}:
            return Response({'detail': _('Unsupported export format.')}, status=status.HTTP_400_BAD_REQUEST)

        if export_format == 'csv':
            rows = _serialize_queryset(
                InsightProjectSerializer,
                queryset.iterator(),
                context=self.get_serializer_context(),
            )
            return stream_csv_response('insightzen-projects.csv', columns, rows)

        rows = _serialize_queryset(
            InsightProjectSerializer,
            queryset.iterator(),
            context=self.get_serializer_context(),
        )
        return build_xlsx_response('insightzen-projects.xlsx', 'Projects', columns, rows)


def _merge_meta(original: dict[str, Any] | None, updates: dict[str, Any] | None) -> dict[str, Any]:
    base = dict(original or {})
    if not updates:
        return base
    for key, value in updates.items():
        base[key] = value
    return base


class QuotaSchemeViewSet(viewsets.ModelViewSet):
    serializer_class = QuotaSchemeSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = InsightZenPagination

    def get_queryset(self):
        queryset = QuotaScheme.objects.select_related('project', 'created_by')
        params = self.request.query_params
        project_id = params.get('project') or params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        status_param = params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        search = params.get('q')
        if search:
            queryset = queryset.filter(name__icontains=search)
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return queryset.order_by('-is_default', '-priority', 'name')
        allowed_ids = _panel_accessible_projects(user, QUOTA_PANEL_KEY)
        if not allowed_ids:
            return queryset.none()
        return queryset.filter(project_id__in=allowed_ids).order_by('-is_default', '-priority', 'name')

    def _ensure_manager(self, project_id: int) -> None:
        membership = _ensure_panel_permission(self.request.user, project_id, QUOTA_PANEL_KEY)
        if self.request.user.is_superuser or self.request.user.is_staff:
            return
        if membership and membership.role in MANAGER_ROLES:
            return
        raise PermissionDenied(_('Only project managers can manage quota schemes.'))

    def perform_create(self, serializer):
        project = serializer.validated_data['project']
        self._ensure_manager(project.id)
        scheme = serializer.save(created_by=self.request.user)
        if scheme.is_default:
            scheme.ensure_default()

    def perform_update(self, serializer):
        scheme = serializer.instance
        self._ensure_manager(scheme.project_id)
        if not scheme.can_edit() and 'status' in serializer.validated_data:
            raise ValidationError({'status': _('Only draft schemes can be edited.')})
        updated_scheme = serializer.save()
        if updated_scheme.is_default:
            updated_scheme.ensure_default()

    def destroy(self, request, *args, **kwargs):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        scheme.mark_archived()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='publish')
    def publish(self, request, pk=None):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        serializer = QuotaSchemePublishSerializer(data=request.data or {})
        serializer.is_valid(raise_exception=True)
        is_default = serializer.validated_data.get('is_default')
        if is_default is not None:
            scheme.is_default = bool(is_default)
        scheme.mark_published()
        update_fields = []
        if is_default is not None:
            update_fields.append('is_default')
        if update_fields:
            scheme.save(update_fields=update_fields)
        if scheme.is_default:
            scheme.ensure_default()
        scheme.refresh_from_db()
        return Response(self.get_serializer(scheme).data)

    @action(detail=True, methods=['post'], url_path='archive')
    def archive(self, request, pk=None):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        scheme.mark_archived()
        return Response(self.get_serializer(scheme).data)

    @action(detail=True, methods=['get'], url_path='cells')
    def list_cells(self, request, pk=None):
        scheme = self.get_object()
        _ensure_panel_permission(request.user, scheme.project_id, QUOTA_PANEL_KEY)
        cells = list(scheme.cells.all().order_by('id'))
        policy = scheme.overflow_policy
        complete_param = request.query_params.get('complete')
        if complete_param is not None:
            desired = complete_param.lower() in {'true', '1', 'yes'}
            filtered: list[QuotaCell] = []
            for cell in cells:
                remaining = cell.remaining_slots(policy)
                is_complete = remaining == 0 if remaining is not None else False
                if desired and is_complete:
                    filtered.append(cell)
                elif not desired and not is_complete:
                    filtered.append(cell)
            cells = filtered
        search = request.query_params.get('q')
        if search:
            term = search.lower()
            filtered_cells: list[QuotaCell] = []
            for cell in cells:
                selector_text = json.dumps(cell.selector or {}, ensure_ascii=False).lower()
                label_text = (cell.label or '').lower()
                if term in label_text or term in selector_text:
                    filtered_cells.append(cell)
            cells = filtered_cells
        serializer = QuotaCellSerializer(cells, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='cells/bulk_upsert')
    def bulk_upsert_cells(self, request, pk=None):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        serializer = QuotaCellBulkUpsertSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        payload = serializer.validated_data['cells']
        updated_ids: list[int] = []
        for row in payload:
            if not isinstance(row, dict):
                raise ValidationError(_('Each cell definition must be an object.'))
            selector = row.get('selector')
            if not isinstance(selector, dict):
                selector = {
                    key: value
                    for key, value in row.items()
                    if key not in {'label', 'target', 'soft_cap', 'weight', 'selector'} and value not in (None, '')
                }
            if not selector:
                raise ValidationError(_('Selector information is required for each cell.'))
            try:
                target = int(row.get('target'))
            except (TypeError, ValueError) as exc:
                raise ValidationError(_('Target must be provided as an integer.')) from exc
            soft_cap_value = row.get('soft_cap')
            soft_cap = None
            if soft_cap_value not in (None, ''):
                try:
                    soft_cap = int(soft_cap_value)
                except (TypeError, ValueError) as exc:
                    raise ValidationError(_('Soft cap must be a valid integer.')) from exc
            weight_value = row.get('weight', 1.0)
            try:
                weight = float(weight_value)
            except (TypeError, ValueError) as exc:
                raise ValidationError(_('Weight must be numeric.')) from exc
            cell, _created = QuotaCell.objects.update_or_create(
                scheme=scheme,
                selector=selector,
                defaults={
                    'label': row.get('label', ''),
                    'target': max(target, 0),
                    'soft_cap': None if soft_cap is None else max(soft_cap, 0),
                    'weight': max(weight, 0.0001),
                },
            )
            updated_ids.append(cell.id)
        cells = list(scheme.cells.filter(id__in=updated_ids))
        return Response(QuotaCellSerializer(cells, many=True).data)

    @action(detail=True, methods=['post'], url_path='pool/build')
    def build_pool(self, request, pk=None):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        payload = request.data or {}
        cell_ids = payload.get('cells') or payload.get('cell_ids')
        limit_param = payload.get('limit')
        multiplier_param = payload.get('multiplier')

        cells_qs = scheme.cells.all()
        if cell_ids not in (None, ''):
            if not isinstance(cell_ids, (list, tuple, set)):
                raise ValidationError({'cells': _('Cell identifiers must be provided as a list.')})
            try:
                ids = [int(value) for value in cell_ids]
            except (TypeError, ValueError) as exc:
                raise ValidationError({'cells': _('Invalid cell identifier provided.')}) from exc
            cells_qs = cells_qs.filter(pk__in=ids)

        limit_value: int | None = None
        if limit_param not in (None, ''):
            try:
                limit_value = max(int(limit_param), 0)
            except (TypeError, ValueError) as exc:
                raise ValidationError({'limit': _('Limit must be a positive integer.')}) from exc

        multiplier_value = 5
        if multiplier_param not in (None, ''):
            try:
                multiplier_value = max(int(multiplier_param), 1)
            except (TypeError, ValueError) as exc:
                raise ValidationError({'multiplier': _('Multiplier must be a positive integer.')}) from exc

        cells = list(cells_qs.order_by('id'))
        if not cells:
            raise ValidationError({'cells': _('No quota cells found for the provided criteria.')})

        results: list[dict[str, Any]] = []
        total_inserted = 0
        try:
            for cell in cells:
                inserted = build_sample_pool_for_cell(
                    cell,
                    limit=limit_value,
                    multiplier=multiplier_value,
                )
                results.append({'cell': cell.id, 'inserted': inserted})
                total_inserted += inserted
        except ProgrammingError as exc:
            raise ValidationError({'detail': _('Bank schema is not available: %(error)s') % {'error': str(exc)}})

        return Response({'results': results, 'inserted_total': total_inserted})

    @action(detail=True, methods=['patch'], url_path='cells/(?P<cell_pk>[^/.]+)')
    def update_cell(self, request, pk=None, cell_pk=None):
        scheme = self.get_object()
        self._ensure_manager(scheme.project_id)
        cell = get_object_or_404(QuotaCell, pk=cell_pk, scheme=scheme)
        serializer = QuotaCellUpdateSerializer(cell, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        cell.refresh_from_db()
        return Response(QuotaCellSerializer(cell).data)

    @action(detail=True, methods=['get'], url_path='stats')
    def stats(self, request, pk=None):
        scheme = self.get_object()
        _ensure_panel_permission(request.user, scheme.project_id, QUOTA_PANEL_KEY)
        cells = list(scheme.cells.all())
        policy = scheme.overflow_policy
        target_total = sum(cell.target for cell in cells)
        achieved_total = sum(cell.achieved for cell in cells)
        in_progress_total = sum(cell.in_progress for cell in cells)
        remaining_total = sum((cell.remaining_slots(policy) or 0) for cell in cells)
        by_dimension: dict[str, dict[str, dict[str, int]]] = {}
        for dimension in scheme.dimensions or []:
            key = dimension.get('key')
            if not key:
                continue
            dimension_totals: dict[str, dict[str, int]] = {}
            for cell in cells:
                selector_value = cell.selector.get(key) if isinstance(cell.selector, dict) else None
                value_key = str(selector_value if selector_value is not None else _('Unspecified'))
                entry = dimension_totals.setdefault(
                    value_key,
                    {'target': 0, 'achieved': 0, 'in_progress': 0},
                )
                entry['target'] += cell.target
                entry['achieved'] += cell.achieved
                entry['in_progress'] += cell.in_progress
            by_dimension[key] = dimension_totals
        data = {
            'target_total': target_total,
            'achieved_total': achieved_total,
            'in_progress_total': in_progress_total,
            'remaining_total': remaining_total,
            'by_dimension': by_dimension,
        }
        stats_serializer = QuotaSchemeStatsSerializer(data)
        return Response(stats_serializer.data)


class DialerAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = DialerAssignmentSerializer
    permission_classes = (IsAuthenticated,)
    pagination_class = InsightZenPagination

    def get_queryset(self):
        queryset = (
            DialerAssignment.objects.select_related(
                'project', 'scheme', 'cell', 'sample', 'interviewer', 'interview'
            )
            .order_by('-reserved_at')
        )
        user = self.request.user
        if user.is_staff or user.is_superuser:
            filtered = queryset
        else:
            allowed_project_ids = _panel_accessible_projects(user, TELEPHONE_PANEL_KEY)
            if not allowed_project_ids:
                return queryset.none()
            memberships = _get_memberships_by_project(user)
            manager_projects = {
                project_id
                for project_id, membership in memberships.items()
                if project_id in allowed_project_ids and membership.role in MANAGER_ROLES
            }
            agent_projects = allowed_project_ids - manager_projects
            conditions = []
            if manager_projects:
                conditions.append(Q(project_id__in=manager_projects))
            if agent_projects:
                conditions.append(Q(project_id__in=agent_projects, interviewer=user))
            if not conditions:
                return queryset.none()
            combined = conditions.pop()
            for condition in conditions:
                combined |= condition
            filtered = queryset.filter(combined)
        return self._apply_filters(filtered)

    def _apply_filters(self, queryset):
        params = self.request.query_params
        project_id = params.get('project') or params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        status_param = params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        interviewer_param = params.get('interviewer')
        if interviewer_param:
            queryset = queryset.filter(interviewer_id=interviewer_param)
        search = params.get('q')
        if search:
            queryset = queryset.filter(sample__phone_number__icontains=search)
        return queryset

    def perform_create(self, serializer):
        project = serializer.validated_data['project']
        membership = _ensure_panel_permission(self.request.user, project.id, TELEPHONE_PANEL_KEY)
        if membership and membership.role not in MANAGER_ROLES:
            raise PermissionDenied(_('Only project managers can create assignments.'))
        sample = serializer.validated_data['sample']
        if sample.status != SampleContact.STATUS_AVAILABLE:
            raise ValidationError({'sample': _('Sample is not available for assignment.')})
        interviewer = serializer.validated_data.get('interviewer')
        if interviewer is None:
            raise ValidationError({'interviewer': _('Interviewer is required.')})
        with transaction.atomic():
            SampleContact.objects.filter(pk=sample.pk).update(
                status=SampleContact.STATUS_CLAIMED,
                attempt_count=F('attempt_count') + 1,
                last_attempt_at=timezone.now(),
                interviewer=interviewer,
                used_at=timezone.now(),
            )
            assignment = serializer.save()
            assignment.cell.increment_in_progress()
            assignment.sample.refresh_from_db()
        return assignment

    def perform_update(self, serializer):
        assignment = serializer.instance
        membership = _ensure_panel_permission(self.request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        if membership and membership.role not in MANAGER_ROLES and assignment.interviewer_id != self.request.user.id:
            raise PermissionDenied(_('You cannot modify this assignment.'))
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        assignment = self.get_object()
        membership = _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        if membership and membership.role not in MANAGER_ROLES:
            raise PermissionDenied(_('You cannot cancel this assignment.'))
        assignment.mark_cancelled()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='expire')
    def expire(self, request, pk=None):
        assignment = self.get_object()
        _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        assignment.mark_expired()
        meta_updates = request.data.get('meta')
        if meta_updates:
            assignment.meta = _merge_meta(assignment.meta, meta_updates)
            assignment.save(update_fields=['meta'])
        serializer = self.get_serializer(assignment)
        return Response(serializer.data)

    def _ensure_assignment_actor(self, request, assignment: DialerAssignment) -> None:
        membership = _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        if request.user.is_superuser or request.user.is_staff:
            return
        if membership and membership.role in MANAGER_ROLES:
            return
        if assignment.interviewer_id != request.user.id:
            raise PermissionDenied(_('You cannot update this assignment.'))

    @action(detail=True, methods=['post'], url_path='complete')
    def complete(self, request, pk=None):
        assignment = self.get_object()
        self._ensure_assignment_actor(request, assignment)
        serializer = AssignmentStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        outcome_code = serializer.validated_data.get('outcome_code')
        meta = serializer.validated_data.get('meta')
        assignment.mark_completed(outcome_code)
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            interview = None
        if interview:
            interview.mark_completed(outcome_code)
            if meta:
                interview.meta = _merge_meta(interview.meta, meta)
                interview.save(update_fields=['status', 'end_form', 'outcome_code', 'meta'])
        if meta:
            assignment.meta = _merge_meta(assignment.meta, meta)
            assignment.save(update_fields=['meta'])
        return Response(self.get_serializer(assignment).data)

    @action(detail=True, methods=['post'], url_path='failed')
    def failed(self, request, pk=None):
        assignment = self.get_object()
        self._ensure_assignment_actor(request, assignment)
        serializer = AssignmentFailSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        outcome_code = serializer.validated_data.get('outcome_code') or 'FAIL'
        meta = serializer.validated_data.get('meta')
        reason = serializer.validated_data.get('reason')
        assignment.mark_failed(outcome_code)
        if meta or reason:
            meta_updates = meta or {}
            if reason:
                meta_updates = {**meta_updates, 'failure_reason': reason}
            assignment.meta = _merge_meta(assignment.meta, meta_updates)
            assignment.save(update_fields=['meta'])
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            interview = None
        if interview and (meta or reason):
            interview.mark_completed(outcome_code)
            meta_updates = meta or {}
            if reason:
                meta_updates = {**meta_updates, 'failure_reason': reason}
            interview.meta = _merge_meta(interview.meta, meta_updates)
            interview.save(update_fields=['status', 'end_form', 'outcome_code', 'meta'])
        return Response(self.get_serializer(assignment).data)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        assignment = self.get_object()
        membership = _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        if membership and membership.role not in MANAGER_ROLES and assignment.interviewer_id != request.user.id:
            raise PermissionDenied(_('You cannot cancel this assignment.'))
        assignment.mark_cancelled()
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            interview = None
        meta_updates = request.data.get('meta')
        if meta_updates:
            assignment.meta = _merge_meta(assignment.meta, meta_updates)
            assignment.save(update_fields=['meta'])
            if interview:
                interview.meta = _merge_meta(interview.meta, meta_updates)
                interview.save(update_fields=['status', 'end_form', 'outcome_code', 'meta'])
        return Response(self.get_serializer(assignment).data)

    @action(detail=True, methods=['get'], url_path='sample')
    def sample(self, request, pk=None):
        assignment = self.get_object()
        serializer = SampleContactSerializer(assignment.sample, context={'request': request})
        return Response(serializer.data)


class InterviewViewSet(viewsets.ViewSet):
    permission_classes = (IsAuthenticated,)

    def _get_assignment(self, request, pk: str) -> DialerAssignment:
        assignment = get_object_or_404(
            DialerAssignment.objects.select_related('project', 'interview'),
            pk=pk,
        )
        _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
        return assignment

    def retrieve(self, request, pk=None):
        assignment = self._get_assignment(request, pk)
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            return Response({'detail': _('Interview has not started.')}, status=status.HTTP_404_NOT_FOUND)
        return Response(InterviewSerializer(interview).data)

    @action(detail=True, methods=['post'], url_path='start')
    def start(self, request, pk=None):
        assignment = self._get_assignment(request, pk)
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            interview = Interview.objects.create(assignment=assignment)
        if request.user != assignment.interviewer and not (
            request.user.is_superuser or request.user.is_staff
        ):
            membership = _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
            if membership and membership.role not in MANAGER_ROLES:
                raise PermissionDenied(_('Only the assigned interviewer can start this interview.'))
        interview.mark_in_progress()
        meta = request.data.get('meta')
        if meta:
            interview.meta = _merge_meta(interview.meta, meta)
            interview.save(update_fields=['status', 'start_form', 'meta'])
        return Response(InterviewSerializer(interview).data)

    @action(detail=True, methods=['post'], url_path='complete')
    def complete(self, request, pk=None):
        assignment = self._get_assignment(request, pk)
        if request.user != assignment.interviewer and not (
            request.user.is_superuser or request.user.is_staff
        ):
            membership = _ensure_panel_permission(request.user, assignment.project_id, TELEPHONE_PANEL_KEY)
            if membership and membership.role not in MANAGER_ROLES:
                raise PermissionDenied(_('Only the assigned interviewer can complete this interview.'))
        serializer = InterviewActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        outcome_code = serializer.validated_data.get('outcome_code')
        meta = serializer.validated_data.get('meta')
        assignment.mark_completed(outcome_code)
        try:
            interview = assignment.interview
        except Interview.DoesNotExist:
            interview = None
        if interview:
            interview.mark_completed(outcome_code)
            if meta:
                interview.meta = _merge_meta(interview.meta, meta)
                interview.save(update_fields=['status', 'end_form', 'outcome_code', 'meta'])
        if meta:
            assignment.meta = _merge_meta(assignment.meta, meta)
            assignment.save(update_fields=['meta'])
        return Response(InterviewSerializer(assignment.interview).data)


class CollectionPerformanceSummaryView(CollectionPerformanceBaseView):
    def get(self, request):
        filters, queryset = self.get_filters_and_queryset(request)
        summary = _calculate_collection_summary(
            queryset,
            project_id=filters['project_id'],
            date_from=filters['date_from'],
            date_to=filters['date_to'],
        )
        return Response(summary)


class CollectionPerformanceBarView(CollectionPerformanceBaseView):
    def get(self, request):
        _filters, queryset = self.get_filters_and_queryset(request)
        group_by = request.query_params.get('group_by', 'interviewer')
        metric = request.query_params.get('metric', 'completes')
        try:
            limit = int(request.query_params.get('limit', 30))
        except ValueError as exc:
            raise ValidationError({'limit': _('Limit must be numeric.')}) from exc

        if limit <= 0:
            limit = 30

        metric_key = 'completes' if metric not in {'attempts', 'sr'} else metric

        if group_by == 'interviewer':
            data = _aggregate_by_interviewer(queryset)
            data.sort(key=lambda item: item.get(metric_key, 0), reverse=True)
            return Response([{
                'interviewer_id': item['interviewer_id'],
                'label': item['label'],
                'attempts': item['attempts'],
                'completes': item['completes'],
                'sr': item['sr'],
                'avg_duration_sec': item['avg_duration_sec'],
                'value': item.get(metric_key, 0),
            } for item in data[:limit]])

        if group_by == 'day':
            daily_queryset = queryset.annotate(
                day=TruncDate(Coalesce('end_form', 'start_form')),
            ).values('day').annotate(
                attempts=Count('id'),
                completes=Count('id', filter=Q(outcome_code='COMP')),
            ).order_by('day')
            results: list[dict[str, Any]] = []
            for entry in daily_queryset:
                attempts = int(entry.get('attempts') or 0)
                completes = int(entry.get('completes') or 0)
                sr_value = (completes / attempts) if attempts else 0.0
                value_map = {
                    'attempts': attempts,
                    'completes': completes,
                    'sr': sr_value,
                }
                metric_value = value_map.get(metric_key, completes)
                results.append(
                    {
                        'label': entry['day'].isoformat() if entry['day'] else '',
                        'attempts': attempts,
                        'completes': completes,
                        'sr': sr_value,
                        'value': metric_value,
                    }
                )
            results.sort(key=lambda item: item.get('label'))
            return Response(results[:limit])

        raise ValidationError({'group_by': _('Unsupported grouping for chart.')})


class CollectionPerformancePieView(CollectionPerformanceBaseView):
    def get(self, request):
        _, queryset = self.get_filters_and_queryset(request)
        metric = request.query_params.get('metric', 'completes')
        metric_key = 'completes' if metric not in {'attempts', 'sr'} else metric
        data = _aggregate_by_interviewer(queryset)
        total = sum(item.get(metric_key, 0) or 0 for item in data)
        results = []
        for item in data:
            value = item.get(metric_key, 0) or 0
            share = (value / total) if total else 0.0
            results.append(
                {
                    'interviewer_id': item['interviewer_id'],
                    'label': item['label'],
                    'value': value,
                    'share': share,
                }
            )
        results.sort(key=lambda item: item['value'], reverse=True)
        return Response(results)


class CollectionPerformanceTopView(CollectionPerformanceBaseView):
    def get(self, request):
        _, queryset = self.get_filters_and_queryset(request)
        sort_key = request.query_params.get('sort', 'completes')
        sort_key = sort_key if sort_key in {'completes', 'attempts', 'sr'} else 'completes'
        try:
            limit = int(request.query_params.get('limit', 5))
        except ValueError as exc:
            raise ValidationError({'limit': _('Limit must be numeric.')}) from exc
        if limit <= 0:
            limit = 5

        data = _aggregate_by_interviewer(queryset)
        data.sort(key=lambda item: item.get(sort_key, 0), reverse=True)
        top_rows = []
        for index, item in enumerate(data[:limit], start=1):
            top_rows.append(
                {
                    'rank': index,
                    'interviewer_id': item['interviewer_id'],
                    'label': item['label'],
                    'attempts': item['attempts'],
                    'completes': item['completes'],
                    'sr': item['sr'],
                    'avg_duration_sec': item['avg_duration_sec'],
                }
            )
        return Response(top_rows)


class CollectionPerformanceTableView(CollectionPerformanceBaseView):
    pagination_class = InsightZenPagination

    def get(self, request):
        _filters, queryset = self.get_filters_and_queryset(request)
        ordered = queryset.order_by('-effective_timestamp', '-pk')
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(ordered, request)
        results: list[dict[str, Any]] = []
        for interview in page:
            assignment = interview.assignment
            project = assignment.project
            interviewer = assignment.interviewer
            sample = assignment.sample
            cell = assignment.cell
            profile = getattr(interviewer, 'insight_profile', None) if interviewer else None
            effective_time = interview.end_form or interview.start_form
            results.append(
                {
                    'date': effective_time.date().isoformat() if effective_time else None,
                    'project': project.name if project else None,
                    'project_code': project.code if project else None,
                    'project_id': project.id if project else None,
                    'interviewer': _interviewer_display_name(interviewer) if interviewer else None,
                    'interviewer_id': interviewer.id if interviewer else None,
                    'team': getattr(profile, 'team', '') if profile else '',
                    'phone_number': _mask_phone_number(getattr(sample, 'phone_number', None)),
                    'outcome_code': interview.outcome_code,
                    'start_form': interview.start_form.isoformat() if interview.start_form else None,
                    'end_form': interview.end_form.isoformat() if interview.end_form else None,
                    'duration_sec': _duration_seconds(interview),
                    'call_attempts': getattr(sample, 'attempt_count', None),
                    'cell_id': cell.id if cell else None,
                    'cell_label': cell.label if cell else None,
                    'gender': getattr(sample, 'gender', None),
                    'province_code': getattr(sample, 'province_code', None),
                    'age_band': getattr(sample, 'age_band', None),
                }
            )
        return paginator.get_paginated_response(results)


class CollectionPerformanceExportView(CollectionPerformanceBaseView):
    def post(self, request):
        filters, queryset = self.get_filters_and_queryset(request)
        summary = _calculate_collection_summary(
            queryset,
            project_id=filters['project_id'],
            date_from=filters['date_from'],
            date_to=filters['date_to'],
        )
        interviewer_stats = _aggregate_by_interviewer(queryset)
        sorted_stats = sorted(interviewer_stats, key=lambda item: item['completes'], reverse=True)
        top_stats = sorted(interviewer_stats, key=lambda item: item['completes'], reverse=True)[:5]

        workbook = Workbook(write_only=True)

        summary_sheet = workbook.create_sheet(title='Summary')
        summary_sheet.append(['Project', summary['project']])
        summary_sheet.append(['From', summary['range']['from']])
        summary_sheet.append(['To', summary['range']['to']])
        summary_sheet.append([])
        summary_sheet.append(['Metric', 'Value'])
        summary_sheet.append(['Attempts', summary['totals']['attempts']])
        summary_sheet.append(['Completes', summary['totals']['completes']])
        summary_sheet.append(['Success Rate', summary['totals']['success_rate']])
        summary_sheet.append(['Avg Duration (sec)', summary['totals']['avg_duration_sec']])
        summary_sheet.append([])
        summary_sheet.append(['Day', 'Attempts', 'Completes', 'Success Rate'])
        for entry in summary['by_day']:
            summary_sheet.append(
                [entry['day'], entry['attempts'], entry['completes'], entry['sr']]
            )

        share_sheet = workbook.create_sheet(title='Interviewer Share')
        share_sheet.append(['Interviewer', 'Attempts', 'Completes', 'Success Rate'])
        for item in sorted_stats:
            share_sheet.append(
                [
                    item['label'],
                    item['attempts'],
                    item['completes'],
                    item['sr'],
                ]
            )

        top_sheet = workbook.create_sheet(title='Top 5')
        top_sheet.append(['Rank', 'Interviewer', 'Attempts', 'Completes', 'Success Rate', 'Avg Duration (sec)'])
        for index, item in enumerate(top_stats, start=1):
            top_sheet.append(
                [
                    index,
                    item['label'],
                    item['attempts'],
                    item['completes'],
                    item['sr'],
                    item['avg_duration_sec'],
                ]
            )

        raw_sheet = workbook.create_sheet(title='Raw Data')
        raw_sheet.append(
            [
                'date',
                'project',
                'project_code',
                'project_id',
                'interviewer',
                'interviewer_id',
                'team',
                'phone_number',
                'outcome_code',
                'call_attempts',
                'start_form',
                'end_form',
                'duration_sec',
                'cell_id',
                'cell_label',
                'gender',
                'province_code',
                'age_band',
            ]
        )
        raw_queryset = queryset.order_by('effective_timestamp', 'pk')
        for interview in raw_queryset.iterator():
            assignment = interview.assignment
            project = assignment.project
            interviewer = assignment.interviewer
            sample = assignment.sample
            cell = assignment.cell
            profile = getattr(interviewer, 'insight_profile', None) if interviewer else None
            effective_time = interview.end_form or interview.start_form
            raw_sheet.append(
                [
                    effective_time.date().isoformat() if effective_time else None,
                    project.name if project else None,
                    project.code if project else None,
                    project.id if project else None,
                    _interviewer_display_name(interviewer) if interviewer else None,
                    interviewer.id if interviewer else None,
                    getattr(profile, 'team', '') if profile else '',
                    _mask_phone_number(getattr(sample, 'phone_number', None)),
                    interview.outcome_code,
                    getattr(sample, 'attempt_count', None),
                    interview.start_form.isoformat() if interview.start_form else None,
                    interview.end_form.isoformat() if interview.end_form else None,
                    _duration_seconds(interview),
                    cell.id if cell else None,
                    cell.label if cell else None,
                    getattr(sample, 'gender', None),
                    getattr(sample, 'province_code', None),
                    getattr(sample, 'age_band', None),
                ]
            )

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        response = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="collection_performance.xlsx"'
        return response


class CollectionPerformanceOptionsView(CollectionPerformanceBaseView):
    def get(self, request):
        _, queryset = self.get_filters_and_queryset(request)
        interviewer_stats = _aggregate_by_interviewer(queryset)
        interviewer_stats.sort(key=lambda item: item['label'])
        outcome_codes = (
            queryset.order_by()
            .values_list('outcome_code', flat=True)
            .distinct()
        )
        teams = (
            queryset.order_by()
            .values_list('assignment__interviewer__insight_profile__team', flat=True)
            .distinct()
        )
        return Response(
            {
                'interviewers': [
                    {
                        'id': item['interviewer_id'],
                        'label': item['label'],
                    }
                    for item in interviewer_stats
                ],
                'outcome_codes': [code for code in outcome_codes if code],
                'teams': [team for team in teams if team],
            }
        )


class DialerNextNumberView(APIView):
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        project_id = request.data.get('project')
        interviewer_id = request.data.get('interviewer') or request.user.id
        ttl_minutes = request.data.get('ttl_minutes')
        scheme_param = request.data.get('scheme') or request.data.get('scheme_id')

        if not project_id:
            raise ValidationError({'project': _('Project is required.')})

        try:
            project = InsightProject.objects.get(pk=project_id)
        except InsightProject.DoesNotExist as exc:
            raise ValidationError({'project': _('Project not found.')}) from exc

        _ensure_panel_permission(request.user, project.id, TELEPHONE_PANEL_KEY)

        if not (request.user.is_superuser or request.user.is_staff):
            if request.user.id != interviewer_id:
                membership = _ensure_panel_permission(request.user, project.id, TELEPHONE_PANEL_KEY)
                if membership and membership.role not in MANAGER_ROLES:
                    raise PermissionDenied(_('You can only request assignments for yourself.'))

        try:
            interviewer = User.objects.get(pk=interviewer_id)
        except User.DoesNotExist as exc:
            raise ValidationError({'interviewer': _('Interviewer not found.')}) from exc

        try:
            membership = InsightMembership.objects.get(
                user=interviewer, project=project, is_active=True
            )
        except InsightMembership.DoesNotExist as exc:
            raise ValidationError({'interviewer': _('Interviewer is not part of this project.')}) from exc

        if membership.role not in MANAGER_ROLES and not (
            isinstance(membership.panel_permissions, dict)
            and isinstance(membership.panel_permissions.get('collection'), dict)
            and membership.panel_permissions['collection'].get(TELEPHONE_PANEL_KEY)
        ):
            raise ValidationError({'interviewer': _('Interviewer lacks telephone panel access.')})

        try:
            ttl_value = int(ttl_minutes) if ttl_minutes else 15
        except (TypeError, ValueError) as exc:
            raise ValidationError({'ttl_minutes': _('TTL must be an integer.')}) from exc
        if ttl_value <= 0:
            raise ValidationError({'ttl_minutes': _('TTL must be positive.')})

        scheme_id = None
        if scheme_param not in (None, ''):
            try:
                scheme_id = int(scheme_param)
            except (TypeError, ValueError) as exc:
                raise ValidationError({'scheme': _('Scheme must be a valid identifier.')}) from exc

        try:
            assignment = DialerAssignment.reserve_next(
                project=project,
                interviewer=interviewer,
                ttl_minutes=ttl_value,
                scheme_id=scheme_id,
            )
        except ValueError as exc:
            raise ValidationError({'detail': str(exc)}) from exc
        except LookupError:
            raise ValidationError({'detail': _('No available contacts to assign.')})

        serializer = DialerAssignmentSerializer(assignment, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)
